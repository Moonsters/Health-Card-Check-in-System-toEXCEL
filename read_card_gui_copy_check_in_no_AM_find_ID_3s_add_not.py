import requests
import os.path
import datetime
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook

# 設定服務與檔案路徑
API_URL = "http://localhost:8000"
EXCEL_FILE = r"C:\Users\1\Downloads\讀取健保卡toEXCEL\健保卡資料.xlsx"

class CheckInApp:
    def __init__(self, master):
        self.master = master
        master.title("健保卡報到系統")
        master.geometry("700x570")

        # 主要介面元素
        self.status_label = tk.Label(master, text="請點擊開始以啟動報到", font=("Arial", 14), fg="blue")
        self.status_label.pack(pady=10)
        
        check_in_controls_frame = tk.Frame(master)
        check_in_controls_frame.pack(pady=5)

        self.start_button = tk.Button(check_in_controls_frame, text="開始報到", command=self.start_check_in, font=("Arial", 12), width=15)
        self.start_button.pack(side="left", padx=(0, 5))

        self.stop_button = tk.Button(check_in_controls_frame, text="停止報到", command=self.stop_check_in, font=("Arial", 12), width=15, state=tk.DISABLED)
        self.stop_button.pack(side="left", padx=(5, 5))

        self.checkin_count_var = tk.StringVar(value="今日已報到人數：讀取中...")
        self.checkin_count_label = tk.Label(check_in_controls_frame, textvariable=self.checkin_count_var, font=("Arial", 12, "bold"), fg="darkgreen")
        self.checkin_count_label.pack(side="left", padx=(10, 0))

        # --- 報到序號設定介面 ---
        settings_frame = tk.LabelFrame(master, text="報到序號設定", padx=10, pady=10)
        settings_frame.pack(pady=10)

        self.numbering_mode = tk.StringVar(value="auto")

        auto_radio = tk.Radiobutton(settings_frame, text="自動給號 (YYYYMMDD0001 格式)", variable=self.numbering_mode, value="auto",
                                    command=self.toggle_manual_options, font=("Arial", 10))
        auto_radio.pack(anchor="w")

        manual_radio = tk.Radiobutton(settings_frame, text="手動給號", variable=self.numbering_mode, value="manual",
                                      command=self.toggle_manual_options, font=("Arial", 10))
        manual_radio.pack(anchor="w")

        self.manual_options_frame = tk.Frame(settings_frame)
        self.manual_increment_mode = tk.StringVar(value="fixed")

        fixed_radio = tk.Radiobutton(self.manual_options_frame, text="固定值", variable=self.manual_increment_mode, value="fixed",
                                     font=("Arial", 10), command=self.update_increment_entry_labels)
        fixed_radio.pack(side="left", padx=(20, 5))

        increment_radio = tk.Radiobutton(self.manual_options_frame, text="自動累加", variable=self.manual_increment_mode, value="increment",
                                         font=("Arial", 10), command=self.update_increment_entry_labels)
        increment_radio.pack(side="left", padx=(5, 5))

        self.prefix_label = tk.Label(self.manual_options_frame, text="號碼前綴", font=("Arial", 10))
        self.prefix_entry = tk.Entry(self.manual_options_frame, width=5, font=("Arial", 10))

        self.manual_entry_label = tk.Label(self.manual_options_frame, text="起始號碼", font=("Arial", 10))
        self.manual_entry = tk.Entry(self.manual_options_frame, width=10, font=("Arial", 10))

        self.suffix_label = tk.Label(self.manual_options_frame, text="號碼後綴", font=("Arial", 10))
        self.suffix_entry = tk.Entry(self.manual_options_frame, width=5, font=("Arial", 10))

        self.toggle_manual_options()
        self.manual_increment_counter = None

        # --- 新增功能按鈕區塊 ---
        action_frame = tk.Frame(master)
        action_frame.pack(pady=10)
        
        status_buttons_frame = tk.Frame(action_frame)
        status_buttons_frame.pack(side="left", padx=(0, 5))

        self.overdue_button = tk.Button(status_buttons_frame, text="已逾時未報到", command=self.find_overdue_unregistered, font=("Arial", 10), width=15)
        self.overdue_button.grid(row=0, column=0, padx=2, pady=2)
        
        self.not_overdue_button = tk.Button(status_buttons_frame, text="未逾時未報到", command=self.find_not_overdue_unregistered, font=("Arial", 10), width=15)
        self.not_overdue_button.grid(row=0, column=1, padx=2, pady=2)

        self.all_unregistered_button = tk.Button(status_buttons_frame, text="全日所有未報到", command=self.find_all_unregistered, font=("Arial", 10), width=15)
        self.all_unregistered_button.grid(row=1, column=0, padx=2, pady=2)

        self.today_checkin_button = tk.Button(status_buttons_frame, text="今日已報到", command=self.find_today_checkedin, font=("Arial", 10), width=15)
        self.today_checkin_button.grid(row=1, column=1, padx=2, pady=2)

        search_label_frame = tk.LabelFrame(action_frame, text="資料查詢", padx=10, pady=5)
        search_label_frame.pack(side="left", padx=(5, 0))
        self.search_entry = tk.Entry(search_label_frame, width=20, font=("Arial", 10))
        self.search_entry.pack(side="left")
        self.search_entry.bind("<KeyRelease>", self.search_records)

        # --- 資料顯示區塊 ---
        self.data_text_box = tk.Text(master, wrap="word", height=10, width=50, font=("Arial", 12), state=tk.DISABLED)
        self.data_text_box.pack(padx=20, pady=10)

        self.master.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.is_checking_in = False
        self.check_in_loop_id = None
        
        self.wb = self.load_excel_file()
        if self.wb:
            self.update_checkin_count()
            self.start_count_update_loop()
        else:
            messagebox.showerror("啟動失敗", "無法開啟 Excel 檔案，請檢查路徑和檔案是否存在。")
            self.master.destroy()

    def load_excel_file(self):
        """檢查並載入 Excel 檔案，若不存在則建立"""
        if not os.path.exists(EXCEL_FILE):
            try:
                wb = Workbook()
                ws_checkin = wb.active
                ws_checkin.title = "健保卡資料"
                ws_checkin.append(["報到序號", "姓名", "身分證字號", "性別", "出生日期", "卡號", "發卡日期", "報到時間"])
                
                ws_app = wb.create_sheet("預約名單")
                ws_app.append(["預約日期", "預約時間段", "姓名", "身分證字號", "性別", "出生日期"])
                
                wb.save(EXCEL_FILE)
                return wb
            except Exception as e:
                print(f"建立 Excel 檔案失敗：{e}")
                return None
        else:
            try:
                return load_workbook(EXCEL_FILE)
            except Exception as e:
                print(f"載入 Excel 檔案失敗：{e}")
                return None

    def toggle_manual_options(self):
        if self.numbering_mode.get() == "manual":
            self.manual_options_frame.pack(anchor="w", pady=(0, 5))
            self.update_increment_entry_labels()
        else:
            self.manual_options_frame.pack_forget()
        self.manual_increment_counter = None

    def update_increment_entry_labels(self):
        self.prefix_label.pack_forget()
        self.prefix_entry.pack_forget()
        self.manual_entry_label.pack_forget()
        self.manual_entry.pack_forget()
        self.suffix_label.pack_forget()
        self.suffix_entry.pack_forget()
        mode = self.manual_increment_mode.get()
        if mode == "increment":
            self.prefix_label.pack(side="left", padx=(5, 5))
            self.prefix_entry.pack(side="left", padx=(0, 5))
            self.manual_entry_label.config(text="起始號碼")
            self.manual_entry_label.pack(side="left", padx=(5, 5))
            self.manual_entry.pack(side="left")
            self.suffix_label.pack(side="left", padx=(5, 5))
            self.suffix_entry.pack(side="left", padx=(0, 5))
        elif mode == "fixed":
            self.prefix_label.pack(side="left", padx=(5, 5))
            self.prefix_entry.pack(side="left", padx=(0, 5))
            self.manual_entry_label.config(text="固定號碼")
            self.manual_entry_label.pack(side="left", padx=(5, 5))
            self.manual_entry.pack(side="left")
            self.suffix_label.pack(side="left", padx=(5, 5))
            self.suffix_entry.pack(side="left", padx=(0, 5))

    def on_closing(self):
        self.is_checking_in = False
        if self.check_in_loop_id:
            self.master.after_cancel(self.check_in_loop_id)
        self.master.destroy()

    def start_check_in(self):
        if self.is_checking_in:
            return
        self.is_checking_in = True
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.status_label.config(text="請放入健保卡... 正在等待讀取...", fg="black")
        self.check_card_and_run()

    def stop_check_in(self):
        self.is_checking_in = False
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.status_label.config(text="已停止報到，請點擊開始以重新啟動", fg="red")
        if self.check_in_loop_id:
            self.master.after_cancel(self.check_in_loop_id)
            self.check_in_loop_id = None

    def check_card_and_run(self):
        if not self.is_checking_in:
            return

        try:
            response = requests.get(f"{API_URL}/")
            response.raise_for_status()
            readers_data = response.json()
            card_data = None
            for reader in readers_data:
                if reader.get('full_name'):
                    card_data = reader
                    break

            if card_data:
                is_allowed, reason = self.check_if_allowed(card_data.get('id_no'))
                
                if is_allowed:
                    self.process_card_data(card_data)
                elif reason == "not_in_list":
                    if messagebox.askyesno("詢問", f"人員 ({card_data.get('full_name')}) 未在預約名單中。\n是否要將其加入名單並進行報到？"):
                        self.add_to_appointment_list(card_data)
                        self.process_card_data(card_data)
                    else:
                        self.status_label.config(text="已取消報到，請取出健保卡。", fg="red")
                else: # reason == "time_mismatch"
                    self.status_label.config(text="報到失敗：不在預約時間段內。", fg="red")
            
            if card_data:
                self.stop_check_in()
                self.status_label.config(text="報到完成！請取卡... 3秒後自動重新啟動...", fg="green")
                self.check_in_loop_id = self.master.after(3000, self.start_check_in)
            elif self.is_checking_in:
                self.check_in_loop_id = self.master.after(1000, self.check_card_and_run)

        except requests.exceptions.RequestException:
            self.status_label.config(text="錯誤：無法連接服務，請確認 tw-nhi-icc-service.exe 已啟動", fg="red")
            self.stop_check_in()
        except Exception as e:
            messagebox.showerror("錯誤", f"發生未預期的錯誤: {e}")
            self.stop_check_in()

    def check_if_allowed(self, id_no):
        """檢查身分證字號是否在預約名單內，且在約定時間段內"""
        try:
            appointment_sheet = self.wb["預約名單"]
            records = list(appointment_sheet.iter_rows(values_only=True))
            
            id_found_in_list = False
            for row in records[1:]:
                if len(row) < 4 or not row[3]: continue
                
                appointment_date_str = str(row[0]).strip()
                appointment_time_range_str = str(row[1]).strip()
                
                if str(row[3]).strip() == str(id_no).strip():
                    id_found_in_list = True
                    try:
                        current_date = datetime.date.today().strftime('%Y-%m-%d')
                        if current_date != appointment_date_str:
                            continue

                        start_time_str, end_time_str = appointment_time_range_str.split('-')
                        start_time = datetime.datetime.strptime(f"{current_date} {start_time_str}", '%Y-%m-%d %H:%M').time()
                        end_time = datetime.datetime.strptime(f"{current_date} {end_time_str}", '%Y-%m-%d %H:%M').time()
                        
                        current_time = datetime.datetime.now().time()
                        
                        if start_time <= current_time <= end_time:
                            return True, "allowed"
                        else:
                            return False, "time_mismatch"
                    except (ValueError, IndexError):
                        continue
            
            if not id_found_in_list:
                return False, "not_in_list"
            else:
                return False, "time_mismatch"

        except Exception as e:
            messagebox.showerror("錯誤", f"讀取預約名單時發生錯誤：\n{e}")
            return False, "error"

    def add_to_appointment_list(self, card_data):
        """將人員資料加入到預約名單中"""
        try:
            appointment_sheet = self.wb["預約名單"]
            
            current_date = datetime.datetime.now().strftime('%Y-%m-%d')
            current_hour = datetime.datetime.now().hour
            time_period = f"{current_hour:02d}:00-{(current_hour + 1):02d}:00"
            
            new_row = [
                current_date,
                time_period,
                card_data.get('full_name'),
                card_data.get('id_no'),
                card_data.get('sex'),
                card_data.get('birth_date')
            ]
            
            appointment_sheet.append(new_row)
            self.wb.save(EXCEL_FILE)
            print(f"偵錯：人員 {card_data.get('full_name')} 已成功加入預約名單。")
            messagebox.showinfo("成功", f"人員 ({card_data.get('full_name')}) 已成功加入預約名單。")
        except Exception as e:
            messagebox.showerror("錯誤", f"寫入預約名單失敗：\n{e}")

    def process_card_data(self, card_data):
        current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        data_to_write = [
            card_data.get('full_name'),
            card_data.get('id_no'),
            card_data.get('sex'),
            card_data.get('birth_date'),
            card_data.get('card_no'),
            card_data.get('issue_date'),
            current_time
        ]
        self.status_label.config(text=f"讀取成功！歡迎 {card_data.get('full_name')}", fg="green")
        saved_data = self.save_to_excel(data_to_write)
        if saved_data:
            self.data_text_box.config(state=tk.NORMAL)
            self.data_text_box.delete("1.0", tk.END)
            
            # --- 修正顯示格式 ---
            birth_date_display = saved_data[4].split(" ")[0] if saved_data[4] else ""
            issue_date_display = saved_data[6].split(" ")[0] if saved_data[6] else ""

            self.data_text_box.insert("1.0", f"報到序號: {saved_data[0]}\n姓名: {saved_data[1]}\n身分證號: {saved_data[2]}\n性別: {saved_data[3]}\n出生日期: {birth_date_display}\n卡號: {saved_data[5]}\n發卡日期: {issue_date_display}\n報到時間: {saved_data[7]}")
            self.data_text_box.config(state=tk.DISABLED)
            self.update_checkin_count()

    def save_to_excel(self, data_to_write):
        try:
            checkin_sheet = self.wb["健保卡資料"]
            existing_ids = {str(row[2]).strip() for row in checkin_sheet.iter_rows(values_only=True) if len(row) > 2 and row[2]}

            new_id = str(data_to_write[1]).strip()
            if new_id in existing_ids:
                message = f"此身分證字號 ({new_id}) 已存在於試算表中，可能為重複報到。\n您確定要重複寫入嗎？"
                if not messagebox.askyesno("警告：重複報到", message):
                    self.status_label.config(text="已取消報到，資料未寫入。", fg="red")
                    return None

            numbering_mode = self.numbering_mode.get()
            check_in_number = ""
            if numbering_mode == "auto":
                current_date_str = datetime.datetime.now().strftime('%Y%m%d')
                all_records = list(checkin_sheet.iter_rows(values_only=True))
                last_serial_number = 0
                for row in reversed(all_records):
                    if row and str(row[0]).startswith(current_date_str):
                        try:
                            last_serial_number = int(str(row[0])[8:])
                            break
                        except (ValueError, IndexError):
                            continue
                new_serial_number = last_serial_number + 1
                check_in_number = f"{current_date_str}{new_serial_number:04d}"
            elif numbering_mode == "manual":
                manual_increment_mode = self.manual_increment_mode.get()
                prefix_input = self.prefix_entry.get()
                manual_input = self.manual_entry.get()
                suffix_input = self.suffix_entry.get()
                if manual_increment_mode == "fixed":
                    if not manual_input:
                        messagebox.showerror("輸入錯誤", "手動給號輸入框不能為空！")
                        return None
                    check_in_number = prefix_input + manual_input + suffix_input
                elif manual_increment_mode == "increment":
                    if not manual_input or not manual_input.isdigit():
                        messagebox.showerror("輸入錯誤", "自動累加模式的起始號碼必須為純數字！")
                        return None
                    if self.manual_increment_counter is None:
                        try:
                            self.manual_increment_counter = int(manual_input)
                        except ValueError:
                            messagebox.showerror("輸入錯誤", "自動累加模式的起始號碼必須為純數字！")
                            return None
                    check_in_number = prefix_input + str(self.manual_increment_counter) + suffix_input
                    self.manual_increment_counter += 1
            data_to_write.insert(0, check_in_number)
            checkin_sheet.append(data_to_write)
            self.wb.save(EXCEL_FILE)
            self.status_label.config(text=f"報到成功！序號 {check_in_number} 已寫入。", fg="green")
            return data_to_write
        except Exception as e:
            messagebox.showerror("錯誤", f"寫入 Excel 失敗：\n{e}\n\n請檢查檔案是否被其他程式開啟。")
            self.status_label.config(text="報到失敗，請檢查錯誤訊息。", fg="red")
            return None

    def update_checkin_count(self):
        """從 Excel 讀取今日報到人數並更新 UI"""
        try:
            checkin_sheet = self.wb["健保卡資料"]
            today_date_str = datetime.date.today().strftime('%Y-%m-%d')
            
            all_values = list(checkin_sheet.iter_rows(values_only=True))
            
            today_records = [row for row in all_values[1:] if len(row) > 7 and str(row[7]).strip().startswith(today_date_str)]
            
            new_count = len(today_records)
            self.checkin_count_var.set(f"今日已報到人數：{new_count}")

        except Exception as e:
            self.checkin_count_var.set("今日人數：讀取失敗")
            print(f"更新今日人數失敗：{e}")

    def start_count_update_loop(self):
        """每 5 秒自動更新報到人數"""
        self.update_checkin_count()
        self.master.after(5000, self.start_count_update_loop)

    def search_records(self, event=None):
        """搜尋 Excel 檔案中的紀錄"""
        search_term = self.search_entry.get().strip().lower()
        if not search_term:
            self.data_text_box.config(state=tk.NORMAL)
            self.data_text_box.delete("1.0", tk.END)
            self.data_text_box.config(state=tk.DISABLED)
            return

        try:
            checkin_sheet = self.wb["健保卡資料"]
            all_records = list(checkin_sheet.iter_rows(values_only=True))
            headers = all_records[0] if all_records else []
            data_rows = all_records[1:] if all_records else []
            found_records = []
            
            for row in data_rows:
                name = str(row[1]).lower() if len(row) > 1 and row[1] else ""
                id_no = str(row[2]).lower() if len(row) > 2 and row[2] else ""
                if search_term in name or search_term in id_no:
                    found_records.append(row)

            self.data_text_box.config(state=tk.NORMAL)
            self.data_text_box.delete("1.0", tk.END)

            if found_records:
                self.data_text_box.insert("1.0", "找到以下紀錄：\n\n")
                for record in found_records:
                    record_str = ""
                    for i, header in enumerate(headers):
                        if i < len(record) and record[i] is not None:
                            record_str += f"{header}: {record[i]}\n"
                    self.data_text_box.insert(tk.END, record_str + "\n" + "-" * 30 + "\n")
            else:
                self.data_text_box.insert("1.0", "查無符合條件的紀錄。")
            self.data_text_box.config(state=tk.DISABLED)

        except Exception as e:
            messagebox.showerror("錯誤", f"搜尋時發生錯誤：\n{e}")

    def find_all_unregistered(self):
        """比對預約名單與報到名單，找出今日所有未報到人員"""
        try:
            appointment_sheet = self.wb["預約名單"]
            appointment_values = list(appointment_sheet.iter_rows(values_only=True))
            
            checkin_sheet = self.wb["健保卡資料"]
            checkin_values = list(checkin_sheet.iter_rows(values_only=True))
            
            # 修正：確保取得的是乾淨的身分證字號集合
            checked_in_ids = {str(row[2]).strip() for row in checkin_values[1:] if len(row) > 2 and row[2]}
            print(f"\n偵錯：已報到人員的身分證字號集合: {checked_in_ids}")
            
            all_unregistered = []
            today_date_str = datetime.date.today().strftime('%Y-%m-%d')

            for row in appointment_values[1:]:
                # 修正：確保欄位存在且有值
                if len(row) < 4 or not row[0] or not row[3]: 
                    continue
                
                appointment_date_str = str(row[0]).split(" ")[0].strip() # 修正：只取日期部分
                id_no = str(row[3]).strip()
                name = str(row[2]).strip()
                appointment_time_range_str = str(row[1]).strip()

                print(f"偵錯：比對預約名單中的 {name}, ID: {id_no}, 日期: {appointment_date_str}")

                if id_no not in checked_in_ids and appointment_date_str == today_date_str:
                    all_unregistered.append({'姓名': name, '身分證字號': id_no, '預約時間段': appointment_time_range_str})

            self.data_text_box.config(state=tk.NORMAL)
            self.data_text_box.delete("1.0", tk.END)
            
            count = len(all_unregistered)
            if count > 0:
                self.data_text_box.insert("1.0", f"--- 今日所有未報到人員 (共 {count} 位) ---\n\n")
                for person in all_unregistered:
                    self.data_text_box.insert(tk.END, f"姓名: {person['姓名']}\n身分證字號: {person['身分證字號']}\n預約時間: {person['預約時間段']}\n" + "-"*30 + "\n")
            else:
                self.data_text_box.insert("1.0", "今日所有未報到人員 (共 0 位)。恭喜！所有預約人員皆已完成報到。")
                
            self.data_text_box.config(state=tk.DISABLED)

        except Exception as e:
            messagebox.showerror("錯誤", f"比對全日未報到名單時發生錯誤：\n{e}")

    def find_overdue_unregistered(self):
        """比對預約名單與報到名單，找出逾時未報到人員"""
        try:
            appointment_sheet = self.wb["預約名單"]
            appointment_values = list(appointment_sheet.iter_rows(values_only=True))
            
            checkin_sheet = self.wb["健保卡資料"]
            checkin_values = list(checkin_sheet.iter_rows(values_only=True))
            
            # 修正：確保取得的是乾淨的身分證字號集合
            checked_in_ids = {str(row[2]).strip() for row in checkin_values[1:] if len(row) > 2 and row[2]}
            
            overdue_unregistered = []
            current_time = datetime.datetime.now().time()
            today_date_str = datetime.date.today().strftime('%Y-%m-%d')

            for row in appointment_values[1:]:
                # 修正：確保欄位存在且有值
                if len(row) < 4 or not row[0] or not row[3]: 
                    continue
                
                appointment_date_str = str(row[0]).split(" ")[0].strip() # 修正：只取日期部分
                appointment_time_range_str = str(row[1]).strip()
                id_no = str(row[3]).strip()
                name = str(row[2]).strip()

                if id_no not in checked_in_ids and appointment_date_str == today_date_str:
                    try:
                        end_time_str = appointment_time_range_str.split('-')[1]
                        end_time = datetime.datetime.strptime(end_time_str, '%H:%M').time()
                        
                        if current_time > end_time:
                            overdue_unregistered.append({'姓名': name, '身分證字號': id_no, '預約時間段': appointment_time_range_str})
                    except (ValueError, IndexError):
                        continue

            self.data_text_box.config(state=tk.NORMAL)
            self.data_text_box.delete("1.0", tk.END)

            count = len(overdue_unregistered)
            if count > 0:
                self.data_text_box.insert("1.0", f"--- 今日已逾時未報到人員 (共 {count} 位) ---\n\n")
                for person in overdue_unregistered:
                    self.data_text_box.insert(tk.END, f"姓名: {person['姓名']}\n身分證字號: {person['身分證字號']}\n預約時間: {person['預約時間段']}\n" + "-"*30 + "\n")
            else:
                self.data_text_box.insert("1.0", f"今日已逾時未報到人員 (共 0 位)。")
                
            self.data_text_box.config(state=tk.DISABLED)

        except Exception as e:
            messagebox.showerror("錯誤", f"比對逾時未報到名單時發生錯誤：\n{e}")

    def find_not_overdue_unregistered(self):
        """比對預約名單與報到名單，找出今日尚未報到人員（非逾時）"""
        try:
            appointment_sheet = self.wb["預約名單"]
            appointment_values = list(appointment_sheet.iter_rows(values_only=True))
            
            checkin_sheet = self.wb["健保卡資料"]
            checkin_values = list(checkin_sheet.iter_rows(values_only=True))
            
            # 修正：確保取得的是乾淨的身分證字號集合
            checked_in_ids = {str(row[2]).strip() for row in checkin_values[1:] if len(row) > 2 and row[2]}
            
            not_overdue_unregistered = []
            current_time = datetime.datetime.now().time()
            today_date_str = datetime.date.today().strftime('%Y-%m-%d')

            for row in appointment_values[1:]:
                # 修正：確保欄位存在且有值
                if len(row) < 4 or not row[0] or not row[3]: 
                    continue
                
                appointment_date_str = str(row[0]).split(" ")[0].strip() # 修正：只取日期部分
                appointment_time_range_str = str(row[1]).strip()
                id_no = str(row[3]).strip()
                name = str(row[2]).strip()

                if id_no not in checked_in_ids and appointment_date_str == today_date_str:
                    try:
                        end_time_str = appointment_time_range_str.split('-')[1]
                        end_time = datetime.datetime.strptime(end_time_str, '%H:%M').time()
                        
                        if current_time <= end_time:
                            not_overdue_unregistered.append({'姓名': name, '身分證字號': id_no, '預約時間段': appointment_time_range_str})
                    except (ValueError, IndexError):
                        not_overdue_unregistered.append({'姓名': name, '身分證字號': id_no, '預約時間段': appointment_time_range_str})

            self.data_text_box.config(state=tk.NORMAL)
            self.data_text_box.delete("1.0", tk.END)

            count = len(not_overdue_unregistered)
            if count > 0:
                self.data_text_box.insert("1.0", f"--- 今日未逾時未報到人員 (共 {count} 位) ---\n\n")
                for person in not_overdue_unregistered:
                    self.data_text_box.insert(tk.END, f"姓名: {person['姓名']}\n身分證字號: {person['身分證字號']}\n預約時間: {person['預約時間段']}\n" + "-"*30 + "\n")
            else:
                self.data_text_box.insert("1.0", f"今日未逾時未報到人員 (共 0 位)。")
                
            self.data_text_box.config(state=tk.DISABLED)

        except Exception as e:
            messagebox.showerror("錯誤", f"比對未逾時未報到名單時發生錯誤：\n{e}")

    def find_today_checkedin(self):
        """列出今日所有已報到人員"""
        try:
            checkin_sheet = self.wb["健保卡資料"]
            all_records = list(checkin_sheet.iter_rows(values_only=True))
            
            today_date_str = datetime.date.today().strftime('%Y-%m-%d')
            checkedin_records = [row for row in all_records[1:] if len(row) > 7 and str(row[7]).strip().startswith(today_date_str)]
            
            self.data_text_box.config(state=tk.NORMAL)
            self.data_text_box.delete("1.0", tk.END)
            
            count = len(checkedin_records)
            if count > 0:
                self.data_text_box.insert("1.0", f"--- 今日已報到人員 (共 {count} 位) ---\n\n")
                for record in checkedin_records:
                    record_str = ""
                    headers = ["報到序號", "姓名", "身分證字號", "性別", "出生日期", "卡號", "發卡日期", "報到時間"]
                    for i, header in enumerate(headers):
                        if i < len(record) and record[i] is not None:
                            value = record[i]
                            # 修正：將日期格式化為只有日期
                            if header in ["出生日期", "發卡日期"] and isinstance(value, datetime.datetime):
                                value = value.strftime('%Y-%m-%d')
                            
                            record_str += f"{header}: {value}\n"
                    self.data_text_box.insert(tk.END, record_str + "-"*30 + "\n")
            else:
                self.data_text_box.insert("1.0", "今日已報到人員 (共 0 位)。")
                
            self.data_text_box.config(state=tk.DISABLED)

        except Exception as e:
            messagebox.showerror("錯誤", f"讀取今日已報到名單時發生錯誤：\n{e}")


def main():
    root = tk.Tk()
    app = CheckInApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()